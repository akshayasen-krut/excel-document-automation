from docxtpl import RichText, DocxTemplate, InlineImage
from openpyxl import load_workbook
from docx.shared import Mm
import os 

output_dir="created_certificate"
os.makedirs(output_dir,exist_ok=True)
os.chdir(output_dir)

wb=load_workbook("C:/Users/akshaya/Downloads/urcode/wshop.xlsx")
ws=wb.active

#headers
for row in ws.iter_rows(min_row=1,max_row=1,values_only=True):
    headers=list(row)

for row in ws.iter_rows(min_row=2,values_only=True):
       name = row[0]
       workshop = row[1]
       mentor = row[2]
       hod = row[3]
       att = row[4]
       ques = row[5]
       test = row[6]

       overall = round((att + ques + test) /3,2)

       doc=DocxTemplate("C:/Users/akshaya/Downloads/Docs/certificate.docx")

       name_rt=RichText()
       name_rt.add(name, font="italic", size=80)
       workshop_rt=RichText()
       workshop_rt.add(workshop, bold=True, size=40)
       mentor_rt=RichText()
       mentor_rt.add(mentor, bold=True, size=40)
       hod_rt=RichText()
       hod_rt.add(hod, bold=True, size=40)
       overall_rt=RichText()
       overall_rt.add(overall, bold=True, size=40)
       myimage=InlineImage(doc,"C:/Users/akshaya/Downloads/urcode/dummy-modified.png",
                           width=Mm(45), height=Mm(39))

       context ={
              "Name" : name_rt,
             "workshop" : workshop_rt,
             "mentor": mentor_rt,
             "hod" : hod_rt,
             "overall" : overall_rt,
             "logo"  : myimage
                }

       
       doc.render(context)
       output=f"{name}_certificate.docx"
       doc.save(output)